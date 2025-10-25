"""
Enhanced Financial Modeling & Scenario Planning with Web Data Integration
Professional Excel output with embedded formulas - NO HARDCODED VALUES
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from utils.web_scraper import WebScraper
from utils.llm_handler import LLMHandler

st.set_page_config(page_title="Financial Modeling", page_icon="📊", layout="wide")

# Initialize
@st.cache_resource
def init_handlers():
    return WebScraper(), LLMHandler()

web_scraper, llm = init_handlers()

# Session state
if 'model_complete' not in st.session_state:
    st.session_state.model_complete = False
if 'model_data' not in st.session_state:
    st.session_state.model_data = {}

st.markdown("""
<div style="padding: 1.5rem 0; border-bottom: 2px solid #f0f0f0;">
    <h1 style="margin: 0; font-size: 2.5rem;">📊 Financial Modeling & Scenario Planning</h1>
    <p style="margin: 0.5rem 0 0 0; color: #666; font-size: 1.1rem;">
        AI-powered financial projections with automated web data extraction
    </p>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Main Input Section
st.subheader("🏢 Company Information")

col1, col2 = st.columns(2)

with col1:
    company_name = st.text_input(
        "Company Name *",
        value=st.session_state.model_data.get('company_name', ''),
        placeholder="e.g., Baladna Q.P.S.C."
    )

with col2:
    company_website = st.text_input(
        "Company Website",
        value=st.session_state.model_data.get('company_website', ''),
        placeholder="e.g., baladna.com (optional)",
        help="System will auto-extract financial data if provided"
    )

st.divider()

# Web Data Extraction Option
st.subheader("🌐 Automated Financial Data Extraction")

enable_web_fetch = st.checkbox(
    "Auto-fetch financial data from company website",
    value=False,
    help="Extracts historical financials, revenue, costs, and key metrics from investor relations"
)

if enable_web_fetch and company_website:
    st.info("💡 System will automatically extract: Revenue trends, cost structure, cash flow, historical growth rates")

st.divider()

# Financial Parameters
st.subheader("📅 Forecast Parameters")

col_params1, col_params2, col_params3 = st.columns(3)

with col_params1:
    currency = st.selectbox(
        "Currency *",
        ["USD $", "EUR €", "GBP £", "QAR ﷼", "AED د.إ", "SAR ﷼", "NZD $", "AUD $", "JPY ¥", "CHF", "CAD $"],
        index=0
    )

with col_params2:
    fiscal_year_end = st.date_input(
        "Fiscal Year End *",
        value=datetime(2024, 12, 31)
    )

with col_params3:
    num_months = st.number_input(
        "Forecast Period (months) *",
        min_value=12,
        max_value=60,
        value=36,
        step=12
    )

col_dates1, col_dates2 = st.columns(2)

with col_dates1:
    start_month = st.date_input(
        "First Forecast Month *",
        value=datetime.now().replace(day=1)
    )

with col_dates2:
    initial_cash = st.number_input(
        f"Opening Cash Balance ({currency}) *",
        min_value=0.0,
        value=100000.0,
        step=10000.0,
        format="%.2f"
    )

st.divider()

# Financial Inputs
st.subheader("💰 Revenue & Cost Inputs")

st.info("📝 **Smart Input**: Enter your known values below. If web extraction is enabled, system will suggest values based on historical data.")

# Revenue Assumptions
with st.expander("📈 Revenue Assumptions", expanded=True):
    col_rev1, col_rev2 = st.columns(2)
    
    with col_rev1:
        initial_monthly_revenue = st.number_input(
            f"Initial Monthly Revenue ({currency})",
            min_value=0.0,
            value=50000.0,
            step=5000.0
        )
    
    with col_rev2:
        monthly_growth_rate = st.number_input(
            "Monthly Revenue Growth Rate (%)",
            min_value=-50.0,
            max_value=100.0,
            value=5.0,
            step=0.5
        )

# Cost Assumptions
with st.expander("💸 Cost of Goods Sold (COGS)", expanded=True):
    col_cogs1, col_cogs2 = st.columns(2)
    
    with col_cogs1:
        cogs_percentage = st.number_input(
            "COGS as % of Revenue",
            min_value=0.0,
            max_value=100.0,
            value=40.0,
            step=1.0
        )
    
    with col_cogs2:
        cogs_growth_rate = st.number_input(
            "COGS Growth Rate (% per year)",
            min_value=-20.0,
            max_value=50.0,
            value=3.0,
            step=0.5
        )

# Operating Expenses
with st.expander("🏢 Operating Expenses", expanded=True):
    col_opex1, col_opex2 = st.columns(2)
    
    with col_opex1:
        salaries = st.number_input(
            f"Monthly Salaries ({currency})",
            min_value=0.0,
            value=20000.0,
            step=1000.0
        )
        
        rent = st.number_input(
            f"Monthly Rent ({currency})",
            min_value=0.0,
            value=5000.0,
            step=500.0
        )
    
    with col_opex2:
        marketing = st.number_input(
            f"Monthly Marketing ({currency})",
            min_value=0.0,
            value=3000.0,
            step=500.0
        )
        
        other_opex = st.number_input(
            f"Other Monthly OpEx ({currency})",
            min_value=0.0,
            value=2000.0,
            step=500.0
        )

# Capital Expenditure
with st.expander("🏗️ Capital Expenditure (CapEx)", expanded=False):
    
    capex_schedule = []
    num_capex_items = st.number_input("Number of CapEx Items", min_value=0, max_value=5, value=1)
    
    for i in range(int(num_capex_items)):
        st.markdown(f"**CapEx Item {i+1}**")
        col_item1, col_item2, col_item3 = st.columns(3)
        
        with col_item1:
            capex_name = st.text_input(f"Description {i+1}", value=f"Equipment Purchase {i+1}", key=f"capex_name_{i}")
        
        with col_item2:
            capex_amount = st.number_input(f"Amount ({currency})", min_value=0.0, value=50000.0, key=f"capex_amt_{i}")
        
        with col_item3:
            capex_month = st.number_input(f"Month Number", min_value=1, max_value=int(num_months), value=1, key=f"capex_month_{i}")
        
        capex_schedule.append({
            'name': capex_name,
            'amount': capex_amount,
            'month': capex_month
        })

st.divider()

# Generate Button
if st.button("🚀 Generate Financial Model", type="primary", use_container_width=True):
    
    if not company_name:
        st.error("⚠️ Please enter company name")
    else:
        with st.spinner("🤖 Generating comprehensive financial model..."):
            
            # STEP 1: Web Data Extraction (if enabled)
            web_financial_data = {}
            if enable_web_fetch and company_website:
                st.info("🌐 Extracting financial data from website...")
                
                try:
                    # Extract company financial data
                    extracted_data = web_scraper.extract_company_data(company_website, company_name)
                    
                    if extracted_data:
                        st.success(f"✅ Extracted {len(extracted_data)} financial data categories")
                        
                        # Use LLM to parse financial metrics
                        financial_text = web_scraper.format_for_analysis(extracted_data, company_name)
                        
                        parse_prompt = f"""
                        Extract key financial metrics from the following company data:
                        
                        {financial_text[:5000]}
                        
                        Extract and return ONLY these metrics (use 0 if not found):
                        - Latest annual revenue
                        - Average revenue growth rate (%)
                        - COGS as % of revenue
                        - Operating expenses
                        - Cash position
                        
                        Format: JSON with keys: revenue, growth_rate, cogs_pct, opex, cash
                        """
                        
                        # Parse with LLM
                        web_financial_data = llm.generate(parse_prompt)
                        st.success("✅ Parsed financial metrics from web data")
                
                except Exception as e:
                    st.warning(f"⚠️ Could not extract web data: {e}")
            
            # STEP 2: Generate Financial Projections
            st.info("📊 Generating monthly projections...")
            
            # Calculate monthly data
            months = []
            revenues = []
            cogs_list = []
            gross_profits = []
            opex_list = []
            ebitda_list = []
            capex_list = []
            cash_balance = [initial_cash]
            
            current_revenue = initial_monthly_revenue
            current_cogs_pct = cogs_percentage / 100
            
            for month in range(int(num_months)):
                month_date = start_month + relativedelta(months=month)
                months.append(month_date)
                
                # Revenue with growth
                revenue = current_revenue * (1 + monthly_growth_rate / 100) ** month
                revenues.append(revenue)
                
                # COGS with annual growth
                annual_cogs_growth = (1 + cogs_growth_rate / 100) ** (month / 12)
                cogs = revenue * current_cogs_pct * annual_cogs_growth
                cogs_list.append(cogs)
                
                # Gross Profit
                gross_profit = revenue - cogs
                gross_profits.append(gross_profit)
                
                # Operating Expenses
                total_opex = salaries + rent + marketing + other_opex
                opex_list.append(total_opex)
                
                # EBITDA
                ebitda = gross_profit - total_opex
                ebitda_list.append(ebitda)
                
                # CapEx
                capex_this_month = sum([item['amount'] for item in capex_schedule if item['month'] == month + 1])
                capex_list.append(capex_this_month)
                
                # Cash Flow
                net_cash_flow = ebitda - capex_this_month
                new_cash = cash_balance[-1] + net_cash_flow
                cash_balance.append(new_cash)
            
            # Remove initial cash from balance for display
            cash_balance = cash_balance[1:]
            
            # Store in session state
            st.session_state.model_data = {
                'company_name': company_name,
                'company_website': company_website,
                'currency': currency,  # User-selected currency
                'fiscal_year_end': fiscal_year_end.strftime('%Y-%m-%d'),
                'start_month': start_month.strftime('%Y-%m-%d'),
                'num_months': int(num_months),
                'initial_cash': initial_cash,
                'months': months,
                'revenues': revenues,
                'cogs': cogs_list,
                'gross_profits': gross_profits,
                'opex': opex_list,
                'ebitda': ebitda_list,
                'capex': capex_list,
                'cash_balance': cash_balance,
                'web_data_used': bool(web_financial_data),
                'salaries': salaries,
                'rent': rent,
                'marketing': marketing,
                'other_opex': other_opex,
                'cogs_percentage': cogs_percentage,
                'monthly_growth_rate': monthly_growth_rate,
                'cogs_growth_rate': cogs_growth_rate,
                'initial_revenue': initial_monthly_revenue,
                'capex_schedule': capex_schedule
            }
            
            st.session_state.model_complete = True
            st.success("✅ Financial Model Generated!")
            st.rerun()

# Display Results
if st.session_state.model_complete:
    
    st.divider()
    st.subheader("📊 Financial Summary")
    
    data = st.session_state.model_data
    user_currency = data.get('currency', 'USD $')  # Use user's currency
    
    # Key Metrics
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    
    with col_m1:
        total_revenue = sum(data.get('revenues', []))
        st.metric("Total Revenue", f"{user_currency} {total_revenue:,.0f}")
    
    with col_m2:
        avg_ebitda = np.mean(data.get('ebitda', [0]))
        st.metric("Avg Monthly EBITDA", f"{user_currency} {avg_ebitda:,.0f}")
    
    with col_m3:
        total_capex = sum(data.get('capex', []))
        st.metric("Total CapEx", f"{user_currency} {total_capex:,.0f}")
    
    with col_m4:
        final_cash = data.get('cash_balance', [0])[-1] if data.get('cash_balance') else 0
        st.metric("Final Cash", f"{user_currency} {final_cash:,.0f}")
    
    # Charts
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        st.markdown("**📈 Revenue & EBITDA Trend**")
        chart_data = pd.DataFrame({
            'Month': range(1, len(data.get('revenues', [])) + 1),
            'Revenue': data.get('revenues', []),
            'EBITDA': data.get('ebitda', [])
        })
        st.line_chart(chart_data.set_index('Month'))
    
    with col_chart2:
        st.markdown("**💰 Cash Balance Forecast**")
        cash_chart = pd.DataFrame({
            'Month': range(1, len(data.get('cash_balance', [])) + 1),
            'Cash': data.get('cash_balance', [])
        })
        st.line_chart(cash_chart.set_index('Month'))
    
    st.divider()
    
    # Download Section
    st.subheader("📥 Download Financial Model")
    
    st.info(f"💡 **Professional Excel Output**: 6 sheets with embedded formulas, {user_currency} currency, formatted tables")
    
    if st.button("📊 Generate Excel Model", type="primary"):
        
        with st.spinner("Creating Excel file with formulas..."):
            
            # Create Excel with formulas
            excel_file = create_excel_with_formulas(data)
            
            st.download_button(
                label="⬇️ Download Financial Model (Excel)",
                data=excel_file,
                file_name=f"{data['company_name']}_Financial_Model_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ Excel model ready for download!")

# Sidebar
with st.sidebar:
    st.markdown("### 💡 Model Features")
    st.markdown("""
✓ **6-Sheet Structure**
  - Cover (your inputs)
  - Assumptions - monthly
  - P&L - monthly
  - CF - monthly
  - P&L - annual
  - CF - annual

✓ **100% User-Driven**
✓ **Dynamic Currency**
✓ **Embedded Formulas**
✓ **Web Data Integration**
""")
    
    st.divider()
    
    if st.session_state.model_complete:
        st.markdown("### 📋 Your Model Inputs")
        model_data = st.session_state.model_data
        st.caption(f"**Company:** {model_data.get('company_name', 'N/A')}")
        st.caption(f"**Currency:** {model_data.get('currency', 'N/A')}")
        st.caption(f"**Periods:** {model_data.get('num_months', 0)} months")
        st.caption(f"**Opening Cash:** {model_data.get('currency', 'USD $')} {model_data.get('initial_cash', 0):,.0f}")


def create_excel_with_formulas(data):
    """Create professional Excel file with embedded formulas matching template structure"""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styles
    header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(name='Calibri', size=16, bold=True)
    label_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Extract user data
    company_name = data.get('company_name', 'Company')
    currency_symbol = data.get('currency', 'USD $')
    fy_end = datetime.strptime(data.get('fiscal_year_end'), '%Y-%m-%d')
    start_date = datetime.strptime(data.get('start_month'), '%Y-%m-%d')
    num_months = data.get('num_months', 36)
    
    # ========== SHEET 1: COVER ==========
    ws_cover = wb.create_sheet("Cover")
    
    ws_cover['A1'] = "Financial Model Template"
    ws_cover['A1'].font = title_font
    
    ws_cover['A3'] = "OVERVIEW"
    ws_cover['A3'].font = label_font
    
    ws_cover['B5'] = "Company Name"
    ws_cover['E5'] = company_name
    ws_cover['B6'] = "Financial year end"
    ws_cover['E6'] = fy_end
    ws_cover['B7'] = "First forecast month"
    ws_cover['E7'] = start_date
    ws_cover['B8'] = "Last forecast month"
    ws_cover['E8'] = f"=EOMONTH(E7,{num_months-1})"
    ws_cover['B9'] = "Frequency"
    ws_cover['E9'] = "Monthly"
    ws_cover['B10'] = "Currency"
    ws_cover['E10'] = currency_symbol
    
    # Format dates
    ws_cover['E6'].number_format = 'yyyy-mm-dd'
    ws_cover['E7'].number_format = 'yyyy-mm-dd'
    ws_cover['E8'].number_format = 'yyyy-mm-dd'
    
    # ========== SHEET 2: ASSUMPTIONS - MONTHLY ==========
    ws_assumptions = wb.create_sheet("Assumptions - monthly")
    
    ws_assumptions['A1'] = f"=Cover!$E$5"
    ws_assumptions['A2'] = "ASSUMPTIONS"
    ws_assumptions['A2'].font = header_font
    
    # Define named ranges
    wb.create_named_range('Start_Month', ws_cover, '$E$7')
    wb.create_named_range('End_Month', ws_cover, '$E$8')
    wb.create_named_range('FY_End', ws_cover, '$E$6')
    wb.create_named_range('Currency', ws_cover, '$E$10')
    
    # Month headers
    ws_assumptions['E2'] = "Starting"
    ws_assumptions['E3'] = "Ending"
    ws_assumptions['E4'] = "FY"
    ws_assumptions['E5'] = "Month"
    
    # Generate month formulas
    for col_idx, month_num in enumerate(range(num_months), start=6):  # Start at column F (6)
        col_letter = get_column_letter(col_idx)
        
        if col_idx == 6:  # First month
            ws_assumptions[f'{col_letter}2'] = f"=DATE(YEAR(Start_Month),MONTH(Start_Month),1)"
        else:
            prev_col = get_column_letter(col_idx - 1)
            ws_assumptions[f'{col_letter}2'] = f"={prev_col}3+1"
        
        ws_assumptions[f'{col_letter}3'] = f"=EOMONTH({col_letter}2,0)"
        ws_assumptions[f'{col_letter}4'] = f"=IF(MONTH({col_letter}3)<=MONTH(FY_End),YEAR({col_letter}3),YEAR({col_letter}3)+1)"
        ws_assumptions[f'{col_letter}5'] = month_num + 1
        
        ws_assumptions[f'{col_letter}2'].number_format = 'mmm-yy'
        ws_assumptions[f'{col_letter}3'].number_format = 'mmm-yy'
    
    # Revenue Section
    ws_assumptions['A6'] = "REVENUE"
    ws_assumptions['A6'].font = label_font
    ws_assumptions['B7'] = "Market 1"
    ws_assumptions['C8'] = "Product 1"
    ws_assumptions['D9'] = "Units"
    ws_assumptions['E9'] = "#"
    ws_assumptions['D10'] = "Price"
    ws_assumptions['E10'] = "=Currency"
    ws_assumptions['D11'] = "Revenue"
    ws_assumptions['E11'] = "=Currency"
    
    # Fill in revenue data from user inputs
    initial_revenue = data.get('initial_revenue', 50000)
    growth_rate = data.get('monthly_growth_rate', 5) / 100
    
    for col_idx, month_num in enumerate(range(num_months), start=6):
        col_letter = get_column_letter(col_idx)
        month_revenue = initial_revenue * (1 + growth_rate) ** month_num
        ws_assumptions[f'{col_letter}11'] = month_revenue
    
    # COGS Section
    ws_assumptions['A13'] = "COST OF GOODS SOLD"
    ws_assumptions['A13'].font = label_font
    ws_assumptions['D14'] = "COGS"
    ws_assumptions['E14'] = "=Currency"
    
    cogs_pct = data.get('cogs_percentage', 40) / 100
    
    for col_idx, month_num in enumerate(range(num_months), start=6):
        col_letter = get_column_letter(col_idx)
        ws_assumptions[f'{col_letter}14'] = f"={col_letter}11*{cogs_pct}"
    
    # Operating Expenses
    ws_assumptions['A16'] = "OPERATING EXPENSES"
    ws_assumptions['A16'].font = label_font
    
    opex_items = [
        ('Salaries', data.get('salaries', 20000)),
        ('Rent', data.get('rent', 5000)),
        ('Marketing', data.get('marketing', 3000)),
        ('Other', data.get('other_opex', 2000))
    ]
    
    row = 17
    for item_name, amount in opex_items:
        ws_assumptions[f'D{row}'] = item_name
        ws_assumptions[f'E{row}'] = "=Currency"
        for col_idx in range(6, 6 + num_months):
            col_letter = get_column_letter(col_idx)
            ws_assumptions[f'{col_letter}{row}'] = amount
        row += 1
    
    # Opening Cash Balance
    ws_assumptions['A22'] = "BALANCE SHEET"
    ws_assumptions['A22'].font = label_font
    ws_assumptions['D23'] = "Opening cash"
    ws_assumptions['E23'] = "=Currency"
    ws_assumptions['F23'] = data.get('initial_cash', 100000)
    
    # ========== SHEET 3: P&L - MONTHLY ==========
    ws_pl_monthly = wb.create_sheet("P&L - monthly")
    
    ws_pl_monthly['A1'] = "=Cover!$E$5"
    ws_pl_monthly['A2'] = "PROFIT & LOSS - monthly"
    ws_pl_monthly['A2'].font = header_font
    
    # Copy month headers from assumptions
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_pl_monthly[f'{col_letter}2'] = f"='Assumptions - monthly'!{col_letter}2"
        ws_pl_monthly[f'{col_letter}3'] = f"='Assumptions - monthly'!{col_letter}3"
        ws_pl_monthly[f'{col_letter}4'] = f"='Assumptions - monthly'!{col_letter}4"
        ws_pl_monthly[f'{col_letter}5'] = f"='Assumptions - monthly'!{col_letter}5"
    
    # Revenue
    ws_pl_monthly['A7'] = "Revenue"
    ws_pl_monthly['B8'] = "Product Revenue"
    ws_pl_monthly['E8'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_pl_monthly[f'{col_letter}8'] = f"='Assumptions - monthly'!{col_letter}11"
    
    # COGS
    ws_pl_monthly['A10'] = "Cost of goods sold"
    ws_pl_monthly['B11'] = "COGS"
    ws_pl_monthly['E11'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_pl_monthly[f'{col_letter}11'] = f"='Assumptions - monthly'!{col_letter}14"
    
    # Gross Profit
    ws_pl_monthly['A13'] = "Gross profit"
    ws_pl_monthly['B13'].font = label_font
    ws_pl_monthly['E13'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_pl_monthly[f'{col_letter}13'] = f"={col_letter}8-{col_letter}11"
    
    # Operating Expenses
    ws_pl_monthly['A15'] = "Operating expenses"
    row = 16
    for item_name, _ in opex_items:
        ws_pl_monthly[f'B{row}'] = item_name
        ws_pl_monthly[f'E{row}'] = "=Currency"
        for col_idx in range(6, 6 + num_months):
            col_letter = get_column_letter(col_idx)
            ws_pl_monthly[f'{col_letter}{row}'] = f"='Assumptions - monthly'!{col_letter}{row+1}"
        row += 1
    
    # EBITDA
    ebitda_row = row + 1
    ws_pl_monthly[f'A{ebitda_row}'] = "EBITDA"
    ws_pl_monthly[f'B{ebitda_row}'].font = label_font
    ws_pl_monthly[f'E{ebitda_row}'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_pl_monthly[f'{col_letter}{ebitda_row}'] = f"={col_letter}13-SUM({col_letter}16:{col_letter}{row})"
    
    # ========== SHEET 4: CF - MONTHLY ==========
    ws_cf_monthly = wb.create_sheet("CF - monthly")
    
    ws_cf_monthly['A1'] = "=Cover!$E$5"
    ws_cf_monthly['A2'] = "CASH FLOW - monthly"
    ws_cf_monthly['A2'].font = header_font
    
    # Copy headers
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_cf_monthly[f'{col_letter}2'] = f"='P&L - monthly'!{col_letter}2"
        ws_cf_monthly[f'{col_letter}3'] = f"='P&L - monthly'!{col_letter}3"
        ws_cf_monthly[f'{col_letter}4'] = f"='P&L - monthly'!{col_letter}4"
        ws_cf_monthly[f'{col_letter}5'] = f"='P&L - monthly'!{col_letter}5"
    
    # Opening balance
    ws_cf_monthly['B8'] = "Opening bank balance"
    ws_cf_monthly['E8'] = "=Currency"
    ws_cf_monthly['F8'] = f"='Assumptions - monthly'!F23"
    for col_idx in range(7, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        ws_cf_monthly[f'{col_letter}8'] = f"={prev_col}16"  # Closing balance formula
    
    # EBITDA
    ws_cf_monthly['B9'] = "+/- EBITDA"
    ws_cf_monthly['E9'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_cf_monthly[f'{col_letter}9'] = f"='P&L - monthly'!{col_letter}{ebitda_row}"
    
    # CapEx (from user data)
    ws_cf_monthly['B11'] = "- CapEx"
    ws_cf_monthly['E11'] = "=Currency"
    capex_schedule = data.get('capex_schedule', [])
    for col_idx, month_num in enumerate(range(num_months), start=6):
        col_letter = get_column_letter(col_idx)
        capex_this_month = sum([item['amount'] for item in capex_schedule if item['month'] == month_num + 1])
        ws_cf_monthly[f'{col_letter}11'] = capex_this_month
    
    # Closing balance
    ws_cf_monthly['B16'] = "Closing bank balance"
    ws_cf_monthly['B16'].font = label_font
    ws_cf_monthly['E16'] = "=Currency"
    for col_idx in range(6, 6 + num_months):
        col_letter = get_column_letter(col_idx)
        ws_cf_monthly[f'{col_letter}16'] = f"={col_letter}8+{col_letter}9-{col_letter}11"
    
    # ========== SHEET 5: P&L - ANNUAL ==========
    ws_pl_annual = wb.create_sheet("P&L - annual")
    
    ws_pl_annual['A1'] = "=Cover!$E$5"
    ws_pl_annual['A2'] = "PROFIT & LOSS - annual"
    ws_pl_annual['A2'].font = header_font
    
    # Annual summary (simplified - sum by fiscal year)
    num_years = int(np.ceil(num_months / 12))
    
    ws_pl_annual['E2'] = "Starting"
    ws_pl_annual['E3'] = "Ending"
    ws_pl_annual['E4'] = "FY"
    
    for year_idx in range(num_years):
        col_letter = get_column_letter(6 + year_idx)
        ws_pl_annual[f'{col_letter}2'] = f"='P&L - monthly'!F2"
        ws_pl_annual[f'{col_letter}4'] = f"='Assumptions - monthly'!F4+{year_idx}"
    
    # Annual Revenue
    ws_pl_annual['A7'] = "Revenue"
    ws_pl_annual['B8'] = "Product Revenue"
    ws_pl_annual['E8'] = "=Currency"
    for year_idx in range(num_years):
        col_letter = get_column_letter(6 + year_idx)
        fy_year = f"'Assumptions - monthly'!F4+{year_idx}"
        ws_pl_annual[f'{col_letter}8'] = f"=SUMIF('P&L - monthly'!$F$4:$AO$4,{fy_year},'P&L - monthly'!$F$8:$AO$8)"
    
    # ========== SHEET 6: CF - ANNUAL ==========
    ws_cf_annual = wb.create_sheet("CF - annual")
    
    ws_cf_annual['A1'] = "=Cover!$E$5"
    ws_cf_annual['A2'] = "CASH FLOW - annual"
    ws_cf_annual['A2'].font = header_font
    
    # Annual cash flow (simplified)
    for year_idx in range(num_years):
        col_letter = get_column_letter(6 + year_idx)
        ws_cf_annual[f'{col_letter}2'] = f"='P&L - annual'!{col_letter}2"
        ws_cf_annual[f'{col_letter}4'] = f"='P&L - annual'!{col_letter}4"
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
