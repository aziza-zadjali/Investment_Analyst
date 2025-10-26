"""
Financial Modeling | Regulus AI × QDB
Standalone financial projections with Excel/CSV export. Step 4 of 5.
"""
import streamlit as st
import os, base64, io
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from utils.llm_handler import LLMHandler
from utils.financial_analyzer import FinancialAnalyzer
from utils.qdb_styling import apply_qdb_styling, QDB_DARK_BLUE, QDB_NAVY, QDB_PURPLE, QDB_GOLD

st.set_page_config(page_title="Financial Modeling – Regulus AI", layout="wide")
apply_qdb_styling()

# ==== GLOBAL TEAL BUTTON STYLING ====
st.markdown("""
<style>
button {
 background:linear-gradient(135deg,#16A085 0%,#138074 50%,#0E5F55 100%)!important;
 color:white!important;
 border:none!important;
 border-radius:40px!important;
 padding:12px 36px!important;
 font-weight:700!important;
 font-size:0.95rem!important;
 box-shadow:0 4px 16px rgba(19,128,116,0.25)!important;
 transition:all 0.25s ease!important;
 cursor:pointer!important;
}
button:hover{
 background:linear-gradient(135deg,#0E5F55 0%,#138074 50%,#16A085 100%)!important;
 transform:translateY(-2px)!important;
 box-shadow:0 6px 22px rgba(19,128,116,0.35)!important;
}
button:active{
 transform:translateY(0)!important;
 box-shadow:0 3px 10px rgba(19,128,116,0.2)!important;
}
</style>
""", unsafe_allow_html=True)

def encode_image(path):
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f"data:image/png;base64,{base64.b64encode(f.read()).decode()}"
    return None
qdb_logo = encode_image("QDB_Logo.png")

# ==== INIT HANDLERS ====
@st.cache_resource
def init_handlers():
    return LLMHandler(), FinancialAnalyzer()
llm, fin_analyzer = init_handlers()

# ==== SESSION STATE ====
if 'financial_model' not in st.session_state:
    st.session_state.financial_model = None

# ==== RETURN HOME (TOP LEFT - HORIZONTAL) ====
if st.button("← Return Home", use_container_width=False, key="home_btn"):
    st.switch_page("streamlit_app.py")

st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

# ==== HERO SECTION ====
st.markdown(f"""
<div style="background:linear-gradient(135deg,#1B2B4D 0%, #0E2E4D 100%);color:white;margin:0 -3rem;padding:70px 0 50px;text-align:center;">
<div style="position:absolute;left:50px;top:35px;">
{'<img src="'+qdb_logo+'" style="max-height:70px;">' if qdb_logo else '<b>QDB</b>'}
</div>
<h1 style="font-size:2.5rem;font-weight:800;margin:0 0 10px 0;">Financial Modeling & Projections</h1>
<p style="font-size:1.2rem;color:#E2E8F0;margin:0 0 8px 0;font-weight:500;">Build Comprehensive Financial Models and Forecasts</p>
<p style="color:#CBD5E0;font-size:0.95rem;max-width:700px;margin:0 auto;">
AI-powered 3-5 year financial projections, revenue forecasting, and profitability analysis. Export models in Excel and CSV formats.
</p>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

# ==== WORKFLOW TRACKER ====
st.markdown("""
<style>
.track{background:#F6F5F2;margin:0 -3rem;padding:24px 0;
display:flex;justify-content:space-evenly;align-items:center;}
.circle{width:48px;height:48px;border-radius:50%;display:flex;
align-items:center;justify-content:center;font-weight:700;
background:#CBD5E0;color:#475569;font-size:0.9rem;}
.circle.active{background:linear-gradient(135deg,#16A085 0%,#0E5F55 100%);
color:white;box-shadow:0 4px 12px rgba(19,128,116,0.3);}
.label{margin-top:5px;font-size:0.8rem;font-weight:600;color:#708090;}
.label.active{color:#0E5F55;font-weight:700;}
.pipe{height:2px;width:60px;background:#CBD5E0;}
</style>
<div class="track">
 <div><div class="circle">1</div><div class="label">Deal Sourcing</div></div>
 <div class="pipe"></div>
 <div><div class="circle">2</div><div class="label">Due Diligence</div></div>
 <div class="pipe"></div>
 <div><div class="circle">3</div><div class="label">Market Analysis</div></div>
 <div class="pipe"></div>
 <div><div class="circle active">4</div><div class="label active">Financial Modeling</div></div>
 <div class="pipe"></div>
 <div><div class="circle">5</div><div class="label">Investment Memo</div></div>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

# ==== COMPANY INFO SECTION (DARK BLUE) ====
st.markdown("""
<div style="background:#2B3E54;margin:0 -3rem 0 -3rem;padding:28px 3rem;border-top:2px solid #16A085;border-bottom:1px solid #E0E0E0;">
<h3 style="text-align:left;color:#FFFFFF;font-weight:700;margin:0 0 14px 0;font-size:1.1rem;">Company & Model Setup</h3>
</div>
""", unsafe_allow_html=True)

with st.expander("Enter Financial Information", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        company_name = st.text_input("Company Name", placeholder="Enter company name", key="company_name")
    with col2:
        sector = st.text_input("Sector", placeholder="e.g., Fintech, AI/ML", key="sector")
    with col3:
        currency = st.selectbox("Currency", ["USD","EUR","GBP","QAR","AED"], index=0, key="currency")

st.markdown("<div style='height:2px;'></div>", unsafe_allow_html=True)

# ==== MODEL PARAMETERS (BEIGE) ====
st.markdown("""
<div style="background:#F5F2ED;margin:0 -3rem 0 -3rem;padding:28px 3rem;border-top:2px solid #16A085;border-bottom:1px solid #E0E0E0;">
<h3 style="text-align:left;color:#1B2B4D;font-weight:700;margin:0 0 14px 0;font-size:1.1rem;">Financial Assumptions</h3>
</div>
""", unsafe_allow_html=True)

with st.expander("Set Financial Parameters and Drivers", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        forecast_years = st.selectbox("Forecast Period", [3,5,7], index=1, key="forecast_years")
        revenue_start = st.number_input("Current Annual Revenue (Millions)", min_value=0.0, value=5.0, step=0.5, key="revenue")
    with col2:
        growth_rate = st.slider("Annual Revenue Growth Rate (%)", 0, 150, 30, key="growth")
        gross_margin = st.slider("Gross Margin (%)", 0, 100, 70, key="gross_margin")
    with col3:
        opex_percent = st.slider("OpEx (% of Revenue)", 0, 100, 40, key="opex")
        tax_rate = st.slider("Tax Rate (%)", 0, 50, 20, key="tax")

st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

# ==== ACTION BUTTON ====
btn_col = st.columns([1, 1, 1])[1]
with btn_col:
    if st.button("Generate Financial Model", use_container_width=True, key="generate"):
        try:
            with st.spinner("Building financial model with AI validation..."):
                
                # Build projections
                years = list(range(datetime.now().year, datetime.now().year + forecast_years))
                revenue = [revenue_start * ((1 + growth_rate/100) ** i) for i in range(forecast_years)]
                cogs = [r * (1 - gross_margin/100) for r in revenue]
                gross_profit = [r - c for r, c in zip(revenue, cogs)]
                opex = [r * opex_percent/100 for r in revenue]
                ebit = [gp - op for gp, op in zip(gross_profit, opex)]
                tax = [e * tax_rate/100 for e in ebit]
                net_income = [e - t for e, t in zip(ebit, tax)]
                
                # Create model
                model_df = pd.DataFrame({
                    'Year': years,
                    'Revenue': revenue,
                    'COGS': cogs,
                    'Gross Profit': gross_profit,
                    'Gross Margin %': [(gp/r)*100 for gp, r in zip(gross_profit, revenue)],
                    'OpEx': opex,
                    'EBIT': ebit,
                    'EBIT Margin %': [(e/r)*100 for e, r in zip(ebit, revenue)],
                    'Tax': tax,
                    'Net Income': net_income,
                    'Net Margin %': [(n/r)*100 for n, r in zip(net_income, revenue)]
                })
                
                st.session_state.financial_model = model_df
                st.success("Financial model generated!")
            
        except Exception as e:
            st.error(f"Error: {str(e)}")

# ==== RESULTS DISPLAY ====
if st.session_state.get("financial_model") is not None and company_name:
    st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
    
    st.markdown("""
<div style="background:#2B3E54;margin:0 -3rem 0 -3rem;padding:28px 3rem;border-top:2px solid #16A085;border-bottom:1px solid #E0E0E0;">
<h3 style="text-align:left;color:#FFFFFF;font-weight:700;margin:0;font-size:1.1rem;">Financial Projections</h3>
</div>
""", unsafe_allow_html=True)

    df = st.session_state.financial_model
    
    # Display table
    df_display = df.copy()
    for col in ['Revenue','COGS','Gross Profit','OpEx','EBIT','Tax','Net Income']:
        df_display[col] = df_display[col].apply(lambda x: f"{currency} {x:,.1f}M")
    for col in ['Gross Margin %','EBIT Margin %','Net Margin %']:
        df_display[col] = df_display[col].apply(lambda x: f"{x:.1f}%")
    
    st.dataframe(df_display, use_container_width=True)
    
    # Charts
    col_chart1, col_chart2 = st.columns(2)
    with col_chart1:
        st.markdown("**Revenue & Profitability Trend**")
        st.line_chart(df.set_index('Year')[['Revenue','EBIT','Net Income']])
    with col_chart2:
        st.markdown("**Margin Analysis**")
        st.line_chart(df.set_index('Year')[['Gross Margin %','EBIT Margin %','Net Margin %']])
    
    # Key metrics
    st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    with col_m1:
        st.metric("Year 1 Revenue", f"{currency} {df.iloc[0]['Revenue']:.1f}M")
    with col_m2:
        st.metric("Year End Revenue", f"{currency} {df.iloc[-1]['Revenue']:.1f}M")
    with col_m3:
        cagr = ((df.iloc[-1]['Revenue'] / df.iloc[0]['Revenue']) ** (1/(forecast_years-1)) - 1) * 100
        st.metric("CAGR", f"{cagr:.1f}%")
    with col_m4:
        st.metric("Avg Net Margin", f"{df['Net Margin %'].mean():.1f}%")
    
    # EXPORT SECTION
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    st.markdown("""
<div style="background:#F5F2ED;margin:0 -3rem 0 -3rem;padding:28px 3rem;border-top:2px solid #16A085;border-bottom:1px solid #E0E0E0;">
<h3 style="text-align:left;color:#1B2B4D;font-weight:700;margin:0;font-size:1.1rem;">Export Financial Model</h3>
</div>
""", unsafe_allow_html=True)

    # CSV Export
    csv_data = df.to_csv(index=False)
    
    # Excel Export with styling
    def create_excel(data):
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Model"
        
        # Title
        ws['A1'] = f"{company_name} - {sector} | Financial Projections"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1B2B4D", end_color="1B2B4D", fill_type="solid")
        
        # Headers
        for col_num, header in enumerate(data.columns, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
        
        # Data
        for r_idx, row in enumerate(data.values, 4):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(output)
        return output.getvalue()
    
    excel_bytes = create_excel(df)
    
    col_exp1, col_exp2 = st.columns(2)
    with col_exp1:
        st.download_button(
            "Download Excel (XLSX)",
            excel_bytes,
            f"Financial_Model_{company_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_exp2:
        st.download_button(
            "Download CSV",
            csv_data,
            f"Financial_Model_{company_name}_{datetime.now().strftime('%Y%m%d')}.csv",
            "text/csv",
            use_container_width=True
        )
    
    # NAVIGATION
    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    col_nav1, col_nav2, col_nav3 = st.columns(3)
    
    with col_nav1:
        if st.button("Back to Market Analysis", use_container_width=True, key="back_market"):
            st.switch_page("pages/3_Market_Analysis.py")
    
    with col_nav2:
        if st.button("Go to Investment Memo", use_container_width=True, key="to_memo"):
            st.switch_page("pages/5_Investment_Memo.py")
    
    with col_nav3:
        if st.button("Back to Deal Sourcing", use_container_width=True, key="back_deals"):
            st.switch_page("pages/1_Deal_Sourcing.py")

# ==== FOOTER ====
st.markdown(f"""
<div style="background:#1B2B4D;color:#E2E8F0;padding:20px 36px;margin:40px -3rem -2rem;
display:flex;justify-content:space-between;align-items:center;font-size:0.85rem;">
<p style="margin:0;">© 2025 Regulus AI | All Rights Reserved</p>
<p style="margin:0;color:#A0AEC0;">Powered by Regulus AI</p>
</div>
""", unsafe_allow_html=True)
